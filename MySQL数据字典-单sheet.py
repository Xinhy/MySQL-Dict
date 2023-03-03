# -*- coding: utf-8 -*-

from MySQLUtil import MySQLUtil
from ExcelUtil import ExcelUtil
from openpyxl.styles import Font,Side,Alignment,Border
from openpyxl.utils import get_column_letter


def get_tables(host,port,user,passwd,db):
    db = MySQLUtil(host,port,user,passwd,db)
    result = db.fetchall('SHOW TABLE STATUS')
    
    tables = []
    fields = {}
    for i in result:
        t = {}
        t['name'] = i['Name']
        t['engine'] = i['Engine']
        t['collation'] = i['Collation']
        t['comment'] = i['Comment']
        
        tables.append(t)
        fields[i['Name']] = db.fetchall('SHOW FULL FIELDS FROM ' + i['Name'])
        
    return tables,fields

    

def generate_file(filename,db,tables,fields):
    
    xls = ExcelUtil(filename + '.xlsx')
    # 初始化目录
    if not xls.has_sheet('目录'):
        xls.rename_sheet('Sheet','目录')
        xls.write_data('目录',[('数据库','表名','表中文名','表结构')])
    
    tseq = 1
    for i in tables: 
        # 遍历单表字段       
        data = []
        seq = 1
        for j in fields[i['name']]:
            data.append(('',seq,j['Field'],j['Type'],j['Null'],j['Key'],j['Extra'],j['Comment']))
            seq += 1
        # 处理单表标题    
        headers = [
            ('',),
            ('','表{} {} （{}）'.format(tseq,i['name'],i['comment']),),
            ('','序号','字段名','类型（长度）','Null','Key','其它','备注')
        ]
        tseq += 1
        cur_rows = xls.get_row_num(db) if xls.has_sheet(db) else 0
        xls.write_data(db,headers + data)
        
        
        # 设置格式
        ws = xls.workbook[db]   
        ws.merge_cells('B{}:H{}'.format(cur_rows + 2,cur_rows + 2)) 
        for row in range(cur_rows + 1 ,cur_rows + seq + 3):
            ws.row_dimensions[row].height = 20
        ws.column_dimensions[get_column_letter(2)].width = 10
        ws.column_dimensions[get_column_letter(3)].width = 20
        ws.column_dimensions[get_column_letter(4)].width = 20
        ws.column_dimensions[get_column_letter(5)].width = 12
        ws.column_dimensions[get_column_letter(6)].width = 12
        ws.column_dimensions[get_column_letter(7)].width = 18
        ws.column_dimensions[get_column_letter(8)].width = 30
        
            
        for row in range(cur_rows + 1,cur_rows + seq + 3):
            for col in range(1,9):
                ws.cell(row,col).font = Font(name='宋体',size=10)
                ws.cell(row,col).alignment = Alignment(horizontal='center',vertical='center')
                if row > cur_rows + 1 and col > 1:
                    thick = Side(style='thin')
                    ws.cell(row,col).border = Border(left=thick,right=thick,top=thick,bottom=thick)
        
                    
        xls.write_data('目录',[(db,i['name'],i['comment'],'查看')])
        row2 = xls.get_row_num('目录')
        xls.workbook['目录'].cell(row2,4).hyperlink = "#'{}'!B{}".format(db,cur_rows + 2)
        # xls.workbook[i['comment']].cell(1,1).hyperlink = "#目录!B{}".format(row2)

        
    xls.save()
    
    # 调整目录格式
    ws = xls.workbook['目录']
    ws.column_dimensions[get_column_letter(1)].width = 20
    ws.column_dimensions[get_column_letter(2)].width = 40
    ws.column_dimensions[get_column_letter(3)].width = 40
    ws.column_dimensions[get_column_letter(4)].width = 10
    
    for row in range(1,xls.get_row_num('目录') + 1):
        for col in range(1,5):
            ws.row_dimensions[row].height = 20
            ws.cell(row,col).font = Font(name='宋体',size=12)
            thick = Side(style='thin')
            ws.cell(row,col).border = Border(left=thick,right=thick,top=thick,bottom=thick)
            if col == 4:
                ws.cell(row,col).alignment = Alignment(horizontal='center',vertical='center')
            else:
                ws.cell(row,col).alignment = Alignment(vertical='center')
                
    for i in range(1,5):
        ws.cell(1,i).alignment = Alignment(horizontal='center',vertical='center')
                
    xls.save()
        

    
if __name__ == '__main__':
    host = '127.0.0.1'      #数据库ip
    port = 3306             #端口号
    user = 'root'           #用户名，如为多个数据库，此用户需有数据库的访问权限
    passwd = 'abc123'       #密码   
    database = ['db1','db2']    #导出的数据库，可配置多个
    filename = '数据字典2'       #生成的数据字典文件名
    
    for db in database:
        tables,fields = get_tables(host,port,user,passwd,db)
        generate_file(filename,db,tables,fields)


