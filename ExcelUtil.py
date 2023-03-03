# -*- encoding: utf-8 -*-


from openpyxl import load_workbook
from openpyxl import Workbook
import os


class ExcelUtil:
    """处理 Excel的通用功能封装"""

    def __init__(self, file_path: str):
        """初始化，文件存在则打开，不存在则创建"""
        self.file_path = file_path
        if not os.path.exists(self.file_path):
            self.workbook = Workbook(self.file_path)
            self.save()
        self.workbook = load_workbook(self.file_path)
        
    def save(self):
        """保存文档"""
        self.workbook.save(self.file_path)

       
    def create_sheet(self,sheet_name: str):
        """创建sheet"""
        self.workbook.create_sheet(sheet_name)
        
    def remove_sheet(self, sheet_name: str):
        """删除名为sheet_name的sheet"""
        self.workbook.remove_sheet(self.workbook[sheet_name])
            
    def rename_sheet(self,old_name: str,new_name: str):
        """更改sheet名称"""
        self.workbook[old_name].title = new_name

    def has_sheet(self, sheet_name: str) -> bool:
        """判断是否存在名为sheet_name的sheet"""
        return sheet_name in self.workbook.sheetnames        

    def read_sheet(self, sheet):
        """读取sheet"""
        return self.workbook.worksheets[sheet].values

    def write_data(self, sheet_name: str, content: list):
        """写入数据"""
        if not self.has_sheet(sheet_name):
            self.workbook.create_sheet(sheet_name)

        for i in content:
            self.workbook[sheet_name].append(i)
    
    def get_row_num(self,sheet_name):
        return self.workbook[sheet_name].max_row
    
    def get_col_num(self,sheet_name):
        return self.workbook[sheet_name].max_column




    